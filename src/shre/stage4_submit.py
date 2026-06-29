import os
import csv

def _generate_reasoning(cand):
    """
    Generates specific, detailed, non-hallucinated reasoning from actual candidate data,
    phrased as a realistic and professional evaluation sentence.
    """
    profile = cand.get('profile', {})
    signals = cand.get('redrob_signals', {})
    career = cand.get('career_history', [])

    title = profile.get('current_title', 'Engineer')
    years = profile.get('years_of_experience', 0)
    company = profile.get('current_company', 'N/A')

    # Start sentence
    intro = f"{title} with {years:.1f} years of experience at {company}"

    # Skills
    skills = cand.get('skills', [])
    ml_skills = [s.get('name') for s in skills if any(x in s.get('name', '').lower() for x in ['ml', 'python', 'pytorch', 'tensorflow', 'nlp'])][:2]
    vector_skills = [s.get('name') for s in skills if any(x in s.get('name', '').lower() for x in ['vector', 'milvus', 'pinecone', 'faiss', 'rag'])][:2]
    backend_skills = [s.get('name') for s in skills if any(x in s.get('name', '').lower() for x in ['spark', 'airflow', 'kafka', 'sql'])][:2]

    focus = ""
    if vector_skills:
        focus = f", specializing in vector search and RAG ({', '.join(vector_skills)})"
    elif ml_skills:
        focus = f", specializing in applied ML ({', '.join(ml_skills)})"

    infra = ""
    if backend_skills:
        infra = f"; experienced in backend infrastructure ({', '.join(backend_skills)})"

    # Trajectory
    growth = ""
    if len(career) >= 2:
        senior_roles = sum(1 for j in career if any(x in j.get('title', '').lower() for x in ['senior', 'lead', 'principal']))
        if senior_roles >= 1:
            growth = ". Shown solid seniority growth in past roles"

    # Behavioral
    behavior = []
    github_score = signals.get('github_activity_score', -1)
    resp_rate = signals.get('recruiter_response_rate', 0)

    if github_score > 70:
        behavior.append("strong GitHub activity")
    elif github_score > 30:
        behavior.append("active GitHub presence")

    if resp_rate > 0.8:
        behavior.append("very high recruiter responsiveness")
    elif resp_rate > 0.5:
        behavior.append("good recruiter responsiveness")

    behavior_str = ""
    if behavior:
        behavior_str = "; features " + " and ".join(behavior)

    # Semantic fit (Feature 1) - grounded in the multi-vector fusion score.
    semantic = cand.get('_semantic', {})
    fusion = semantic.get('semantic_fusion_score')
    skills_sim = semantic.get('semantic_skills_sim', 0)
    traj_sim = semantic.get('semantic_traj_sim', 0)
    semantic_str = ""
    if fusion is not None:
        if fusion >= 0.7:
            best = "skills" if skills_sim >= traj_sim else "experience trajectory"
            semantic_str = f"; strong semantic alignment to the JD (esp. {best})"
        elif fusion >= 0.55:
            semantic_str = "; moderate semantic fit to the JD"

    # Behavioral signals (Feature 4) beyond the raw github/response above.
    beh = cand.get('_behavioral', {})
    beh_bits = []
    if beh.get('demand_score', 0) >= 0.6:
        beh_bits.append("high recruiter demand")
    if beh.get('oss_score', 0) >= 0.6:
        beh_bits.append("strong open-source footprint")
    if beh.get('reliability_score', 0) >= 0.7:
        beh_bits.append("reliable follow-through")
    beh_str = ("; " + ", ".join(beh_bits)) if beh_bits else ""

    # Risk Flags + anomaly notes (Feature 3).
    notice = signals.get('notice_period_days', 0)
    open_to_work = signals.get('open_to_work_flag', False)

    flags = []
    if notice > 90:
        flags.append(f"a long {notice}-day notice period")
    if not open_to_work:
        flags.append("not actively seeking new roles")

    flags_str = ""
    if flags:
        flags_str = ". Note: candidate has " + " and is ".join(flags)

    # Concise data-quality note; the full flag list lives in submission_detailed.csv.
    anomaly = cand.get('_anomaly', {})
    anomaly_flags = anomaly.get('flags', []) if isinstance(anomaly, dict) else []
    anomaly_str = ""
    if anomaly_flags:
        kinds = sorted({f.split(':', 1)[0] for f in anomaly_flags})
        anomaly_str = f". Minor data-quality flag ({', '.join(kinds)})"

    # Assemble
    reasoning = (f"{intro}{focus}{infra}{growth}{semantic_str}"
                 f"{behavior_str}{beh_str}{flags_str}{anomaly_str}.")
    reasoning = reasoning.replace("..", ".").replace(" .", ".").replace("  ", " ").strip()

    # Trim cleanly at a word boundary (no mid-word cuts) and end with a period.
    limit = 300
    if len(reasoning) > limit:
        reasoning = reasoning[:limit].rsplit(' ', 1)[0].rstrip(' ,;[') + '.'
    return reasoning

def export_submission(candidates, scores, out_path):
    """
    Sorts candidates by score, breaks ties by candidate_id, and writes top 100 to CSV.
    """
    import os
    dir_name = os.path.dirname(out_path)
    if dir_name:
        os.makedirs(dir_name, exist_ok=True)
    # Combine candidate ID, score, and the raw candidate object
    scored_candidates = []
    for cand, score in zip(candidates, scores):
        scored_candidates.append({
            'candidate_id': cand.get('candidate_id'),
            'score': float(score),
            'raw': cand
        })
        
    # Sort by score descending, then candidate_id ascending
    scored_candidates.sort(key=lambda x: (x['score'], x['candidate_id']), reverse=True)
    
    # Need to reverse candidate_id sort because reverse=True sorted both descending.
    # To do it correctly:
    scored_candidates.sort(key=lambda x: (-x['score'], x['candidate_id']))
    
    top_100 = scored_candidates[:100]
    
    print(f"Writing top 100 to {out_path}...")
    with open(out_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
        writer.writerow(['candidate_id', 'rank', 'score', 'reasoning'])
        
        for rank, item in enumerate(top_100, 1):
            cand_id = item['candidate_id']
            score = item['score']
            reasoning = _generate_reasoning(item['raw'])
            
            writer.writerow([cand_id, rank, score, reasoning])
            
    # Save the full rankings as requested in the plan
    full_out_path = os.path.join(os.path.dirname(out_path), 'rankings_full.csv')
    print(f"Writing all {len(scored_candidates)} viable candidates to {full_out_path}...")
    with open(full_out_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
        writer.writerow(['candidate_id', 'rank', 'score', 'reasoning'])
        for rank, item in enumerate(scored_candidates, 1):
            writer.writerow([item['candidate_id'], rank, item['score'], _generate_reasoning(item['raw'])])

    # Detailed top-100 view exposing the new enrichment signals (for the UI /
    # analysis). The canonical submission.csv above stays a clean 4 columns.
    detailed_path = os.path.join(os.path.dirname(out_path), 'submission_detailed.csv')
    with open(detailed_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
        writer.writerow(['candidate_id', 'rank', 'score', 'semantic_fit',
                         'behavioral_score', 'anomaly_score', 'anomaly_flags', 'reasoning'])
        for rank, item in enumerate(top_100, 1):
            raw = item['raw']
            sem = raw.get('_semantic', {}) or {}
            beh = raw.get('_behavioral', {}) or {}
            anom = raw.get('_anomaly', {}) or {}
            writer.writerow([
                item['candidate_id'], rank, round(item['score'], 4),
                round(sem.get('semantic_fusion_score', 0.0), 4),
                round(beh.get('behavioral_composite', 0.0), 4),
                round(anom.get('anomaly_score', 0.0), 4),
                '; '.join(anom.get('flags', [])),
                _generate_reasoning(raw),
            ])

    # Ranked XLSX deliverable (submission requirement). Mirrors the CSVs in a
    # formatted workbook: "Top 100" + "Full Rankings" sheets. Never breaks the
    # pipeline — if the Excel writer is unavailable the CSVs already exist.
    xlsx_path = os.path.splitext(out_path)[0] + '.xlsx'
    try:
        _write_xlsx(xlsx_path, top_100, scored_candidates)
        print(f"Writing ranked XLSX to {xlsx_path}...")
    except Exception as e:  # noqa: BLE001 - CSVs are the guaranteed output
        print(f"[stage4] XLSX export skipped ({type(e).__name__}: {e}); "
              f"CSV outputs are unaffected.")

    print("Done!")


def _detailed_rows(items):
    """Build (rank, candidate_id, score, semantic, behavioral, anomaly, flags,
    reasoning) tuples for a list of scored items."""
    rows = []
    for rank, item in enumerate(items, 1):
        raw = item['raw']
        sem = raw.get('_semantic', {}) or {}
        beh = raw.get('_behavioral', {}) or {}
        anom = raw.get('_anomaly', {}) or {}
        rows.append((
            rank,
            item['candidate_id'],
            round(float(item['score']), 4),
            round(sem.get('semantic_fusion_score', 0.0), 4),
            round(beh.get('behavioral_composite', 0.0), 4),
            round(anom.get('anomaly_score', 0.0), 4),
            '; '.join(anom.get('flags', [])),
            _generate_reasoning(raw),
        ))
    return rows


def _write_xlsx(xlsx_path, top_100, scored_candidates):
    """Write a professionally-formatted ranked workbook (openpyxl).

    Sheets: 'Top 100' (recommended shortlist), 'Full Rankings' (all viable),
    and 'Summary' (run statistics). Includes a title block, styled header,
    banded rows, borders, a red→amber→green colour scale on the score column,
    number formats, frozen header and an auto-filter.
    """
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter

    headers = ['rank', 'candidate_id', 'score', 'semantic_fit',
               'behavioral_score', 'anomaly_score', 'anomaly_flags', 'reasoning']
    widths = [6, 16, 9, 12, 16, 13, 34, 95]
    # 0-based indices of numeric columns that get a 4-dp number format.
    num_cols = {2, 3, 4, 5}

    INDIGO = '4F46E5'
    INDIGO_DK = '3730A3'
    BAND = 'EEF2FF'
    WHITE = 'FFFFFF'

    title_font = Font(bold=True, size=16, color=WHITE)
    sub_font = Font(italic=True, size=9, color='E0E7FF')
    header_font = Font(bold=True, color=WHITE, size=11)
    header_fill = PatternFill('solid', fgColor=INDIGO)
    title_fill = PatternFill('solid', fgColor=INDIGO_DK)
    band_fill = PatternFill('solid', fgColor=BAND)
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center')
    wrap = Alignment(vertical='top', wrap_text=True)
    thin = Side(style='thin', color='C7D2FE')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ncol = len(headers)
    last_col = get_column_letter(ncol)
    stamp = datetime.now().strftime('%Y-%m-%d %H:%M')

    wb = Workbook()

    def _build(ws, items, title):
        # --- title block (rows 1-2) ---
        ws.merge_cells(f'A1:{last_col}1')
        t = ws['A1']
        t.value = f'  Redrob SHRE — {title}'
        t.font, t.fill, t.alignment = title_font, title_fill, left
        ws.row_dimensions[1].height = 28

        ws.merge_cells(f'A2:{last_col}2')
        s = ws['A2']
        s.value = (f'  Founding Senior AI Engineer · {len(items)} candidates · '
                   f'generated {stamp} · score = fused ensemble + LambdaMART LTR (0–1)')
        s.font, s.fill, s.alignment = sub_font, title_fill, left
        ws.row_dimensions[2].height = 16

        # --- header (row 3) ---
        hdr = 3
        for c, name in enumerate(headers, 1):
            cell = ws.cell(row=hdr, column=c, value=name)
            cell.font, cell.fill = header_font, header_fill
            cell.alignment = center if c != ncol else left
            cell.border = border
        ws.row_dimensions[hdr].height = 20

        # --- data ---
        first_data = hdr + 1
        for i, r in enumerate(items):
            row_idx = first_data + i
            for c, val in enumerate(r, 1):
                cell = ws.cell(row=row_idx, column=c, value=val)
                cell.border = border
                if c == ncol:                      # reasoning
                    cell.alignment = wrap
                elif (c - 1) in num_cols:
                    cell.alignment = center
                    if c != 1:                     # keep rank as plain int
                        cell.number_format = '0.0000'
                else:
                    cell.alignment = center
                if i % 2 == 1:                     # banded rows
                    if not (c == 3):               # leave score col for colour scale
                        cell.fill = band_fill
        last_row = first_data + len(items) - 1

        # --- widths ---
        for c, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w

        # --- colour scale on the score column (C) ---
        if last_row >= first_data:
            ws.conditional_formatting.add(
                f'C{first_data}:C{last_row}',
                ColorScaleRule(
                    start_type='min', start_color='F8696B',      # red (low)
                    mid_type='percentile', mid_value=50, mid_color='FFEB84',  # amber
                    end_type='max', end_color='63BE7B',          # green (high)
                ),
            )
            ws.auto_filter.ref = f'A{hdr}:{last_col}{last_row}'

        ws.freeze_panes = f'A{first_data}'           # keep title+header visible

    ws1 = wb.active
    ws1.title = 'Top 100'
    ws1.sheet_properties.tabColor = INDIGO
    _build(ws1, _detailed_rows(top_100), 'Recommended Shortlist (Top 100)')

    ws2 = wb.create_sheet('Full Rankings')
    _build(ws2, _detailed_rows(scored_candidates), 'Full Candidate Rankings')

    _write_summary_sheet(wb, top_100, scored_candidates)

    os.makedirs(os.path.dirname(xlsx_path) or '.', exist_ok=True)
    wb.save(xlsx_path)


def _write_summary_sheet(wb, top_100, scored_candidates):
    """Add a 'Summary' sheet with run statistics."""
    from openpyxl.styles import Font, PatternFill, Alignment

    def _avg(items, key):
        vals = []
        for it in items:
            d = it['raw'].get(key, {}) or {}
            if key == '_semantic':
                vals.append(d.get('semantic_fusion_score', 0.0))
            elif key == '_behavioral':
                vals.append(d.get('behavioral_composite', 0.0))
            elif key == '_anomaly':
                vals.append(d.get('anomaly_score', 0.0))
        return round(sum(vals) / len(vals), 4) if vals else 0.0

    scores = [float(it['score']) for it in scored_candidates] or [0.0]
    top_scores = [float(it['score']) for it in top_100] or [0.0]

    ws = wb.create_sheet('Summary')
    ws.sheet_properties.tabColor = '10B981'
    ws.merge_cells('A1:B1')
    h = ws['A1']
    h.value = 'Run Summary'
    h.font = Font(bold=True, size=14, color='FFFFFF')
    h.fill = PatternFill('solid', fgColor='3730A3')
    h.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 24

    rows = [
        ('Recommended (Top N)', len(top_100)),
        ('Total viable candidates ranked', len(scored_candidates)),
        ('Top score', round(max(top_scores), 4)),
        ('Lowest score in shortlist', round(min(top_scores), 4)),
        ('Mean score (all viable)', round(sum(scores) / len(scores), 4)),
        ('Avg semantic fit (shortlist)', _avg(top_100, '_semantic')),
        ('Avg behavioral score (shortlist)', _avg(top_100, '_behavioral')),
        ('Avg anomaly score (shortlist)', _avg(top_100, '_anomaly')),
    ]
    label_font = Font(bold=True)
    for i, (label, val) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=label).font = label_font
        ws.cell(row=i, column=2, value=val)
    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 16
